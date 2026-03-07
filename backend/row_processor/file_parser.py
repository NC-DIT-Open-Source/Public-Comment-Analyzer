"""File parser module for CSV and XLSX files."""

import csv
from typing import List, Dict
from dataclasses import dataclass
import chardet
from openpyxl import load_workbook


@dataclass
class ParsedFile:
    """Represents a parsed file with headers and rows."""
    headers: List[str]
    rows: List[Dict[str, str]]
    row_count: int


class FileParser:
    """Parser for CSV and XLSX files."""
    
    def parse(self, file_path: str, file_type: str) -> ParsedFile:
        """
        Parse a CSV or XLSX file.
        
        Args:
            file_path: Path to the file
            file_type: File type ('csv' or 'xlsx')
            
        Returns:
            ParsedFile object with headers, rows, and row count
            
        Raises:
            ValueError: If file type is not supported
            FileNotFoundError: If file does not exist
        """
        if file_type.lower() == 'csv':
            return self._parse_csv(file_path)
        elif file_type.lower() in ['xlsx', 'xls']:
            return self._parse_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
    
    def _detect_encoding(self, file_path: str) -> str:
        """
        Detect file encoding using chardet.
        
        Args:
            file_path: Path to the file
            
        Returns:
            Detected encoding string (e.g., 'utf-8', 'latin-1')
        """
        with open(file_path, 'rb') as f:
            raw_data = f.read()
            result = chardet.detect(raw_data)
            encoding = result['encoding']
            # Default to utf-8 if detection fails
            return encoding if encoding else 'utf-8'
    
    def _parse_csv(self, file_path: str) -> ParsedFile:
        """
        Parse CSV file with proper encoding detection.
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            ParsedFile object with headers, rows, and row count
            
        Raises:
            ValueError: If file is empty or has invalid format
            UnicodeDecodeError: If file encoding cannot be determined
        """
        # Try multiple encodings in order of preference
        encodings_to_try = []
        
        # First, try chardet detection
        detected_encoding = self._detect_encoding(file_path)
        if detected_encoding:
            encodings_to_try.append(detected_encoding)
        
        # Add common fallback encodings
        fallback_encodings = ['utf-8', 'latin-1', 'windows-1252', 'iso-8859-1', 'cp1252']
        for enc in fallback_encodings:
            if enc not in encodings_to_try:
                encodings_to_try.append(enc)
        
        last_error = None
        
        for encoding in encodings_to_try:
            try:
                headers = []
                rows = []
                
                with open(file_path, 'r', encoding=encoding, newline='') as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames if reader.fieldnames else []
                    
                    if not headers:
                        raise ValueError("CSV file has no headers")
                    
                    for row_num, row in enumerate(reader, start=2):  # Start at 2 (after header)
                        try:
                            # Convert all values to strings and handle None values
                            row_dict = {key: (str(value) if value is not None else '') 
                                       for key, value in row.items()}
                            rows.append(row_dict)
                        except Exception as e:
                            print(f"Warning: Error parsing CSV row {row_num}: {str(e)}")
                            # Continue with other rows
                            continue
                
                if not rows:
                    print("Warning: CSV file has no data rows")
                
                print(f"Successfully parsed CSV with encoding: {encoding}")
                return ParsedFile(
                    headers=headers,
                    rows=rows,
                    row_count=len(rows)
                )
            
            except UnicodeDecodeError as e:
                print(f"Failed to decode CSV file with encoding {encoding}: {str(e)}")
                last_error = e
                continue  # Try next encoding
            
            except csv.Error as e:
                print(f"ERROR: CSV parsing error with encoding {encoding}: {str(e)}")
                raise ValueError(f"Invalid CSV format: {str(e)}") from e
            
            except Exception as e:
                print(f"ERROR: Unexpected error parsing CSV with encoding {encoding}: {str(e)}")
                raise
        
        # If we get here, all encodings failed
        print(f"ERROR: Failed to decode CSV file with any supported encoding")
        raise ValueError(f"File encoding not supported. Tried: {', '.join(encodings_to_try)}. Please ensure file is properly encoded.") from last_error
    
    def _parse_xlsx(self, file_path: str) -> ParsedFile:
        """
        Parse XLSX file (first worksheet only).
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            ParsedFile object with headers, rows, and row count
            
        Raises:
            ValueError: If file is empty, corrupted, or has invalid format
        """
        try:
            # Load workbook and get first worksheet
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            
            if not workbook.worksheets:
                raise ValueError("XLSX file has no worksheets")
            
            worksheet = workbook.worksheets[0]
            
            # Get all rows as a list
            all_rows = list(worksheet.iter_rows(values_only=True))
            
            if not all_rows:
                workbook.close()
                raise ValueError("XLSX file has no data")
            
            # First row is headers
            headers = [str(cell) if cell is not None else '' for cell in all_rows[0]]
            
            if not any(headers):  # All headers are empty
                workbook.close()
                raise ValueError("XLSX file has no headers")
            
            # Remaining rows are data
            rows = []
            for row_num, row_values in enumerate(all_rows[1:], start=2):  # Start at 2 (after header)
                try:
                    # Create dictionary mapping headers to values
                    row_dict = {}
                    for i, header in enumerate(headers):
                        # Get value at index i, or empty string if index out of range
                        value = row_values[i] if i < len(row_values) else None
                        row_dict[header] = str(value) if value is not None else ''
                    rows.append(row_dict)
                except Exception as e:
                    print(f"Warning: Error parsing XLSX row {row_num}: {str(e)}")
                    # Continue with other rows
                    continue
            
            workbook.close()
            
            if not rows:
                print("Warning: XLSX file has no data rows")
            
            return ParsedFile(
                headers=headers,
                rows=rows,
                row_count=len(rows)
            )
        
        except Exception as e:
            print(f"ERROR: Failed to parse XLSX file: {str(e)}")
            if 'workbook' in locals():
                try:
                    workbook.close()
                except:
                    pass
            
            # Provide user-friendly error message
            if 'corrupt' in str(e).lower() or 'invalid' in str(e).lower():
                raise ValueError(f"XLSX file is corrupted or invalid: {str(e)}") from e
            else:
                raise
