"""File parser module for CSV and XLSX files."""

import csv
from typing import List, Dict
from dataclasses import dataclass
import chardet
from openpyxl import load_workbook


@dataclass
class ParsedFile:
    """Represents a parsed file with headers and rows."""
    headers: List[str]
    rows: List[Dict[str, str]]
    row_count: int


class FileParser:
    """Parser for CSV and XLSX files."""
    
    def parse(self, file_path: str, file_type: str) -> ParsedFile:
        """
        Parse a CSV or XLSX file.
        
        Args:
            file_path: Path to the file
            file_type: File type ('csv' or 'xlsx')
            
        Returns:
            ParsedFile object with headers, rows, and row count
            
        Raises:
            ValueError: If file type is not supported
            FileNotFoundError: If file does not exist
        """
        if file_type.lower() == 'csv':
            return self._parse_csv(file_path)
        elif file_type.lower() in ['xlsx', 'xls']:
            return self._parse_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
    
    def _detect_encoding(self, file_path: str) -> str:
        """
        Detect file encoding using chardet.
        
        Args:
            file_path: Path to the file
            
        Returns:
            Detected encoding string (e.g., 'utf-8', 'latin-1')
        """
        with open(file_path, 'rb') as f:
            raw_data = f.read()
            result = chardet.detect(raw_data)
            encoding = result['encoding']
            # Default to utf-8 if detection fails
            return encoding if encoding else 'utf-8'
    
    def _parse_csv(self, file_path: str) -> ParsedFile:
        """
        Parse CSV file with proper encoding detection.
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            ParsedFile object with headers, rows, and row count
        """
        # Detect encoding
        encoding = self._detect_encoding(file_path)
        
        headers = []
        rows = []
        
        with open(file_path, 'r', encoding=encoding, newline='') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames if reader.fieldnames else []
            
            for row in reader:
                # Convert all values to strings and handle None values
                row_dict = {key: (str(value) if value is not None else '') 
                           for key, value in row.items()}
                rows.append(row_dict)
        
        return ParsedFile(
            headers=headers,
            rows=rows,
            row_count=len(rows)
        )
    
    def _parse_xlsx(self, file_path: str) -> ParsedFile:
        """
        Parse XLSX file (first worksheet only).
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            ParsedFile object with headers, rows, and row count
        """
        # Load workbook and get first worksheet
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        
        # Get all rows as a list
        all_rows = list(worksheet.iter_rows(values_only=True))
        
        if not all_rows:
            return ParsedFile(headers=[], rows=[], row_count=0)
        
        # First row is headers
        headers = [str(cell) if cell is not None else '' for cell in all_rows[0]]
        
        # Remaining rows are data
        rows = []
        for row_values in all_rows[1:]:
            # Create dictionary mapping headers to values
            row_dict = {}
            for i, header in enumerate(headers):
                # Get value at index i, or empty string if index out of range
                value = row_values[i] if i < len(row_values) else None
                row_dict[header] = str(value) if value is not None else ''
            rows.append(row_dict)
        
        workbook.close()
        
        return ParsedFile(
            headers=headers,
            rows=rows,
            row_count=len(rows)
        )
