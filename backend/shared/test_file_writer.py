"""Unit tests for FileWriter class."""

import os
import tempfile
import csv
from openpyxl import load_workbook
import pytest

from backend.shared.file_writer import FileWriter


class TestFileWriter:
    """Test cases for FileWriter class."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.writer = FileWriter()
        self.temp_dir = tempfile.mkdtemp()
    
    def test_write_simple_csv(self):
        """Test writing a simple CSV file."""
        headers = ['Name', 'Age', 'City']
        rows = [
            {'Name': 'Alice', 'Age': '30', 'City': 'New York'},
            {'Name': 'Bob', 'Age': '25', 'City': 'Los Angeles'}
        ]
        
        csv_path = os.path.join(self.temp_dir, 'output.csv')
        self.writer.write(headers, rows, csv_path, 'csv')
        
        # Verify file was created and contains correct data
        assert os.path.exists(csv_path)
        
        with open(csv_path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.DictReader(f)
            result_rows = list(reader)
        
        assert len(result_rows) == 2
        assert result_rows[0] == {'Name': 'Alice', 'Age': '30', 'City': 'New York'}
        assert result_rows[1] == {'Name': 'Bob', 'Age': '25', 'City': 'Los Angeles'}
    
    def test_write_csv_with_special_characters(self):
        """Test writing CSV with special characters (quotes, commas, newlines)."""
        headers = ['Comment', 'Rating']
        rows = [
            {'Comment': 'This is a "quoted" comment', 'Rating': '5'},
            {'Comment': 'Comment with, comma', 'Rating': '4'},
            {'Comment': 'Comment with\nnewline', 'Rating': '3'}
        ]
        
        csv_path = os.path.join(self.temp_dir, 'output_special.csv')
        self.writer.write(headers, rows, csv_path, 'csv')
        
        # Read back and verify proper escaping
        with open(csv_path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.DictReader(f)
            result_rows = list(reader)
        
        assert len(result_rows) == 3
        assert result_rows[0]['Comment'] == 'This is a "quoted" comment'
        assert result_rows[1]['Comment'] == 'Comment with, comma'
        assert result_rows[2]['Comment'] == 'Comment with\nnewline'
    
    def test_write_csv_with_unicode(self):
        """Test writing CSV with unicode characters."""
        headers = ['Name', 'Comment']
        rows = [
            {'Name': 'José', 'Comment': 'Café ☕'},
            {'Name': '李明', 'Comment': '你好世界'}
        ]
        
        csv_path = os.path.join(self.temp_dir, 'output_unicode.csv')
        self.writer.write(headers, rows, csv_path, 'csv')
        
        # Read back and verify unicode preservation
        with open(csv_path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.DictReader(f)
            result_rows = list(reader)
        
        assert result_rows[0]['Name'] == 'José'
        assert result_rows[0]['Comment'] == 'Café ☕'
        assert result_rows[1]['Name'] == '李明'
        assert result_rows[1]['Comment'] == '你好世界'
    
    def test_write_empty_csv(self):
        """Test writing CSV with headers but no data rows."""
        headers = ['Name', 'Age']
        rows = []
        
        csv_path = os.path.join(self.temp_dir, 'output_empty.csv')
        self.writer.write(headers, rows, csv_path, 'csv')
        
        # Verify file has headers but no data
        with open(csv_path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.DictReader(f)
            result_rows = list(reader)
        
        assert len(result_rows) == 0
    
    def test_write_simple_xlsx(self):
        """Test writing a simple XLSX file."""
        headers = ['Name', 'Age', 'City']
        rows = [
            {'Name': 'Alice', 'Age': '30', 'City': 'New York'},
            {'Name': 'Bob', 'Age': '25', 'City': 'Los Angeles'}
        ]
        
        xlsx_path = os.path.join(self.temp_dir, 'output.xlsx')
        self.writer.write(headers, rows, xlsx_path, 'xlsx')
        
        # Verify file was created and contains correct data
        assert os.path.exists(xlsx_path)
        
        workbook = load_workbook(xlsx_path)
        worksheet = workbook.active
        
        # Check headers
        header_row = [cell.value for cell in worksheet[1]]
        assert header_row == ['Name', 'Age', 'City']
        
        # Check data rows
        data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 2
        assert data_rows[0] == ('Alice', '30', 'New York')
        assert data_rows[1] == ('Bob', '25', 'Los Angeles')
        
        workbook.close()
    
    def test_write_xlsx_with_special_characters(self):
        """Test writing XLSX with special characters."""
        headers = ['Comment', 'Rating']
        rows = [
            {'Comment': 'This is a "quoted" comment', 'Rating': '5'},
            {'Comment': 'Comment with, comma', 'Rating': '4'},
            {'Comment': 'Comment with\nnewline', 'Rating': '3'}
        ]
        
        xlsx_path = os.path.join(self.temp_dir, 'output_special.xlsx')
        self.writer.write(headers, rows, xlsx_path, 'xlsx')
        
        # Read back and verify
        workbook = load_workbook(xlsx_path)
        worksheet = workbook.active
        data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        
        assert data_rows[0][0] == 'This is a "quoted" comment'
        assert data_rows[1][0] == 'Comment with, comma'
        assert data_rows[2][0] == 'Comment with\nnewline'
        
        workbook.close()
    
    def test_write_xlsx_with_unicode(self):
        """Test writing XLSX with unicode characters."""
        headers = ['Name', 'Comment']
        rows = [
            {'Name': 'José', 'Comment': 'Café ☕'},
            {'Name': '李明', 'Comment': '你好世界'}
        ]
        
        xlsx_path = os.path.join(self.temp_dir, 'output_unicode.xlsx')
        self.writer.write(headers, rows, xlsx_path, 'xlsx')
        
        # Read back and verify unicode preservation
        workbook = load_workbook(xlsx_path)
        worksheet = workbook.active
        data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        
        assert data_rows[0][0] == 'José'
        assert data_rows[0][1] == 'Café ☕'
        assert data_rows[1][0] == '李明'
        assert data_rows[1][1] == '你好世界'
        
        workbook.close()
    
    def test_write_empty_xlsx(self):
        """Test writing XLSX with headers but no data rows."""
        headers = ['Name', 'Age']
        rows = []
        
        xlsx_path = os.path.join(self.temp_dir, 'output_empty.xlsx')
        self.writer.write(headers, rows, xlsx_path, 'xlsx')
        
        # Verify file has headers but no data
        workbook = load_workbook(xlsx_path)
        worksheet = workbook.active
        
        header_row = [cell.value for cell in worksheet[1]]
        assert header_row == ['Name', 'Age']
        
        # Check that there's only one row (headers)
        assert worksheet.max_row == 1
        
        workbook.close()
    
    def test_write_xlsx_with_empty_values(self):
        """Test writing XLSX with empty string values.
        
        Note: openpyxl stores empty strings as None in cells, which is
        expected behavior for Excel files.
        """
        headers = ['Name', 'Age', 'City']
        rows = [
            {'Name': 'Alice', 'Age': '', 'City': 'New York'},
            {'Name': 'Bob', 'Age': '25', 'City': ''}
        ]
        
        xlsx_path = os.path.join(self.temp_dir, 'output_empty_vals.xlsx')
        self.writer.write(headers, rows, xlsx_path, 'xlsx')
        
        workbook = load_workbook(xlsx_path)
        worksheet = workbook.active
        data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        
        # Empty strings are stored as None in Excel, which is expected
        assert data_rows[0] == ('Alice', None, 'New York')
        assert data_rows[1] == ('Bob', '25', None)
        
        workbook.close()
    
    def test_write_preserves_column_order(self):
        """Test that column order is preserved in output."""
        headers = ['Col3', 'Col1', 'Col2']
        rows = [
            {'Col1': 'A', 'Col2': 'B', 'Col3': 'C'},
            {'Col1': 'D', 'Col2': 'E', 'Col3': 'F'}
        ]
        
        csv_path = os.path.join(self.temp_dir, 'output_order.csv')
        self.writer.write(headers, rows, csv_path, 'csv')
        
        with open(csv_path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.reader(f)
            header_row = next(reader)
            first_data_row = next(reader)
        
        # Verify headers are in specified order
        assert header_row == ['Col3', 'Col1', 'Col2']
        # Verify data follows header order
        assert first_data_row == ['C', 'A', 'B']
    
    def test_write_unsupported_file_type(self):
        """Test that unsupported file types raise ValueError."""
        headers = ['Name']
        rows = [{'Name': 'Test'}]
        
        with pytest.raises(ValueError, match="Unsupported file type"):
            self.writer.write(headers, rows, 'test.txt', 'txt')
    
    def test_write_csv_case_insensitive(self):
        """Test that file type is case insensitive."""
        headers = ['Name']
        rows = [{'Name': 'Test'}]
        
        csv_path = os.path.join(self.temp_dir, 'output_case.csv')
        self.writer.write(headers, rows, csv_path, 'CSV')
        
        assert os.path.exists(csv_path)
    
    def test_write_xlsx_case_insensitive(self):
        """Test that XLSX file type is case insensitive."""
        headers = ['Name']
        rows = [{'Name': 'Test'}]
        
        xlsx_path = os.path.join(self.temp_dir, 'output_case.xlsx')
        self.writer.write(headers, rows, xlsx_path, 'XLSX')
        
        assert os.path.exists(xlsx_path)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
