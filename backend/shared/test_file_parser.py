"""Unit tests for FileParser class."""

import os
import tempfile
import csv
from openpyxl import Workbook
import pytest

from backend.shared.file_parser import FileParser, ParsedFile


class TestFileParser:
    """Test cases for FileParser class."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.parser = FileParser()
        self.temp_dir = tempfile.mkdtemp()
    
    def test_parse_simple_csv(self):
        """Test parsing a simple CSV file."""
        # Create a test CSV file
        csv_path = os.path.join(self.temp_dir, 'test.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Name', 'Age', 'City'])
            writer.writerow(['Alice', '30', 'New York'])
            writer.writerow(['Bob', '25', 'Los Angeles'])
        
        # Parse the file
        result = self.parser.parse(csv_path, 'csv')
        
        # Verify results
        assert result.headers == ['Name', 'Age', 'City']
        assert result.row_count == 2
        assert len(result.rows) == 2
        assert result.rows[0] == {'Name': 'Alice', 'Age': '30', 'City': 'New York'}
        assert result.rows[1] == {'Name': 'Bob', 'Age': '25', 'City': 'Los Angeles'}
    
    def test_parse_csv_with_special_characters(self):
        """Test parsing CSV with special characters (quotes, commas)."""
        csv_path = os.path.join(self.temp_dir, 'test_special.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Comment', 'Rating'])
            writer.writerow(['This is a "quoted" comment', '5'])
            writer.writerow(['Comment with, comma', '4'])
        
        result = self.parser.parse(csv_path, 'csv')
        
        assert result.row_count == 2
        assert result.rows[0]['Comment'] == 'This is a "quoted" comment'
        assert result.rows[1]['Comment'] == 'Comment with, comma'
    
    def test_parse_csv_with_unicode(self):
        """Test parsing CSV with unicode characters."""
        csv_path = os.path.join(self.temp_dir, 'test_unicode.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Name', 'Comment'])
            writer.writerow(['José', 'Café ☕'])
            writer.writerow(['李明', '你好世界'])
        
        result = self.parser.parse(csv_path, 'csv')
        
        assert result.row_count == 2
        assert result.rows[0]['Name'] == 'José'
        assert result.rows[0]['Comment'] == 'Café ☕'
        assert result.rows[1]['Name'] == '李明'
    
    def test_parse_empty_csv(self):
        """Test parsing an empty CSV file."""
        csv_path = os.path.join(self.temp_dir, 'test_empty.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Name', 'Age'])
        
        result = self.parser.parse(csv_path, 'csv')
        
        assert result.headers == ['Name', 'Age']
        assert result.row_count == 0
        assert len(result.rows) == 0
    
    def test_parse_simple_xlsx(self):
        """Test parsing a simple XLSX file."""
        xlsx_path = os.path.join(self.temp_dir, 'test.xlsx')
        
        # Create a test XLSX file
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Name', 'Age', 'City'])
        worksheet.append(['Alice', 30, 'New York'])
        worksheet.append(['Bob', 25, 'Los Angeles'])
        workbook.save(xlsx_path)
        
        # Parse the file
        result = self.parser.parse(xlsx_path, 'xlsx')
        
        # Verify results
        assert result.headers == ['Name', 'Age', 'City']
        assert result.row_count == 2
        assert len(result.rows) == 2
        assert result.rows[0] == {'Name': 'Alice', 'Age': '30', 'City': 'New York'}
        assert result.rows[1] == {'Name': 'Bob', 'Age': '25', 'City': 'Los Angeles'}
    
    def test_parse_xlsx_first_worksheet_only(self):
        """Test that only the first worksheet is parsed."""
        xlsx_path = os.path.join(self.temp_dir, 'test_multi.xlsx')
        
        # Create XLSX with multiple worksheets
        workbook = Workbook()
        
        # First worksheet
        ws1 = workbook.active
        ws1.title = 'Sheet1'
        ws1.append(['Name', 'Value'])
        ws1.append(['First', '1'])
        
        # Second worksheet (should be ignored)
        ws2 = workbook.create_sheet('Sheet2')
        ws2.append(['Other', 'Data'])
        ws2.append(['Second', '2'])
        
        workbook.save(xlsx_path)
        
        # Parse the file
        result = self.parser.parse(xlsx_path, 'xlsx')
        
        # Verify only first worksheet was parsed
        assert result.headers == ['Name', 'Value']
        assert result.row_count == 1
        assert result.rows[0] == {'Name': 'First', 'Value': '1'}
    
    def test_parse_xlsx_with_empty_cells(self):
        """Test parsing XLSX with empty cells."""
        xlsx_path = os.path.join(self.temp_dir, 'test_empty_cells.xlsx')
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Name', 'Age', 'City'])
        worksheet.append(['Alice', None, 'New York'])
        worksheet.append(['Bob', 25, None])
        workbook.save(xlsx_path)
        
        result = self.parser.parse(xlsx_path, 'xlsx')
        
        assert result.row_count == 2
        assert result.rows[0] == {'Name': 'Alice', 'Age': '', 'City': 'New York'}
        assert result.rows[1] == {'Name': 'Bob', 'Age': '25', 'City': ''}
    
    def test_parse_empty_xlsx(self):
        """Test parsing an empty XLSX file."""
        xlsx_path = os.path.join(self.temp_dir, 'test_empty.xlsx')
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Name', 'Age'])
        workbook.save(xlsx_path)
        
        result = self.parser.parse(xlsx_path, 'xlsx')
        
        assert result.headers == ['Name', 'Age']
        assert result.row_count == 0
        assert len(result.rows) == 0
    
    def test_parse_unsupported_file_type(self):
        """Test that unsupported file types raise ValueError."""
        with pytest.raises(ValueError, match="Unsupported file type"):
            self.parser.parse('test.txt', 'txt')
    
    def test_parse_nonexistent_file(self):
        """Test that parsing nonexistent file raises FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            self.parser.parse('nonexistent.csv', 'csv')


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
